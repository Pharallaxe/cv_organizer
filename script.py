import os
import shutil
import openpyxl

# Configurations
excel_file = 'tableur_cv.xlsx'  # Changez cela avec le chemin de votre fichier Excel
cv_folder = 'CV'  # Changez cela avec le chemin où vos CV sont stockés
output_base_folder = 'Entreprises'  # Dossier de sortie pour les entreprises

# Charger le fichier Excel
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

# Création du dossier de base s'il n'existe pas
os.makedirs(output_base_folder, exist_ok=True)

# Lire le fichier Excel
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column, max_row=sheet.max_row):
    entreprise = row[0].value
    if entreprise:
        # Créer un dossier pour l'entreprise
        entreprise_folder = os.path.join(output_base_folder, entreprise)
        os.makedirs(entreprise_folder, exist_ok=True)

        # Ouvrir ou créer un fichier prénoms.txt
        prenoms_file_path = os.path.join(entreprise_folder, 'prenoms.txt')
        with open(prenoms_file_path, 'w') as prenoms_file:
            for col_idx, cell in enumerate(row[1:], start=2):
                prenom = sheet.cell(row=1, column=col_idx).value  # Prénom depuis la première ligne
                if cell.value == 'X':
                    # Écrire le prénom dans prénoms.txt
                    prenoms_file.write(prenom + '\n')
                    
                    # Copier le CV correspondant
                    cv_name = f'{prenom}.pdf'  # Supposant que les CV sont des fichiers PDF
                    cv_source = os.path.join(cv_folder, cv_name)
                    cv_destination = os.path.join(entreprise_folder, cv_name)
                    
                    if os.path.exists(cv_source):
                        shutil.copy(cv_source, cv_destination)
                    else:
                        print(f"CV introuvable pour {prenom}: {cv_source}")

print("Dossiers et fichiers créés avec succès.")
